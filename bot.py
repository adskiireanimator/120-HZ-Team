import telebot
import openpyxl
import datetime
print(datetime.datetime.now())


bot =telebot.TeleBot('5244696447:AAHYv1aiEfA4ZIW1QgMYmyRkRzV1XxUyCkQ')



name =''
age=0
id=''

value=0

@bot.message_handler(content_types=['text'])
def get_text_messages(message):
    wb=openpyxl.load_workbook('varmoney2.xlsx')
    if not (str(message.from_user.id) in wb.sheetnames):
        if message.text=="/reg":
            bot.send_message(message.from_user.id,'как тебя зовут?')
            bot.register_next_step_handler(message,get_name)
        else:
            bot.send_message(message.from_user.id,'write me /reg')
    else:
        ws=wb[str(message.from_user.id)]

        bot.send_message(message.from_user.id,f"{ws['A1'].value}, что ты хочешь? \n /day отчет за день \n  /month отчет за месяц \n /add добавить покупочку")
        bot.register_next_step_handler(message,choose_action)

def day_report(id,date):
    wb=openpyxl.load_workbook('varmoney2.xlsx')
    ws=wb[str(id)]
    report=0
    for i in range(2,ws.max_column+1):
        if str(ws.cell(column=i,row=1).value).split(' ')[0]==date:
            report += int(ws.cell(column=i,row=2).value)
    return str(report)

def month_report(id,date):
    wb=openpyxl.load_workbook('varmoney2.xlsx')
    ws=wb[str(id)]
    report=0
    for i in range(2,ws.max_column+1):
        if str(ws.cell(column=i,row=1).value).split(' ')[0].split('-')[1]==date:
            report += int(ws.cell(column=i,row=2).value)
    return str(report)

def choose_action(message):
    if message.text == '/day':
        report=day_report(message.from_user.id,str(datetime.datetime.now()).split(' ')[0])
        bot.send_message(message.from_user.id,report+f'\n что-нибудь еще?')
    elif message.text == '/month':
        report=month_report(message.from_user.id,str(datetime.datetime.now()).split(' ')[0].split('-')[1])
        bot.send_message(message.from_user.id,report+f'\n что-нибудь еще?')
    elif message.text=='/add':
        bot.send_message(message.from_user.id,'сколько вы потратили?')
        bot.register_next_step_handler(message,purchase_value)
    else:
        bot.send_message(message.from_user.id,"я не знаю таких команд")

def purchase_value(message):
    global value
    try:
        value=int(message.text)
    except Exception:
        bot.send_message(message.from_user.id, 'пиши цифрами, сука')
    id=str(message.from_user.id)
    wb=openpyxl.load_workbook('varmoney2.xlsx')
    ws=wb[id]
    ws.cell(column=ws.max_column+1,row=1).value=str(datetime.datetime.now()).split(' ')[0]
    ws.cell(column=ws.max_column,row=2).value=value
    wb.save('varmoney2.xlsx')




def get_name(message):
    global name
    name=message.text
    bot.send_message(message.from_user.id,"сколько тебе лет")
    bot.register_next_step_handler(message,get_age)

def get_age(message):
    global age
    global name
    global id
    while age==0:
        try:
            age=int(message.text)
        except Exception:
            bot.send_message(message.from_user.id,'пиши цифрами, сука')
    id=str(message.from_user.id)
    wb = openpyxl.load_workbook('./varmoney2.xlsx')
    wb.create_sheet(id)
    sheet=wb.get_sheet_by_name(id)
    sheet['A1']=name
    sheet['A2']=age
    wb.save('varmoney2.xlsx')

    name=''
    age=0
    id=''

    bot.send_message(message.from_user.id,'cпсибо за регистрацию')



bot.polling(none_stop=True,interval=0)
