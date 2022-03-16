import telebot
import openpyxl
import datetime
with open('config.txt', 'r') as s:
    token = s.readline()
bot = telebot.TeleBot(token)
name = ''
age = 0
id_user = ''
category = ''
shop = ''
value = 0


@bot.message_handler(content_types=['text'])
def get_text_messages(message):
    # Основная фунция, которая обрабатывает разные сообщения и распределяет дальнейшие пути развития
    # (15 значущих строчек кода, границу проходит)
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    if not (str(message.from_user.id) in wb.sheetnames):
        if message.text == "/reg":
            bot.send_message(message.from_user.id, 'как тебя зовут?')
            bot.register_next_step_handler(message, get_name)
        else:
            bot.send_message(message.from_user.id, 'напиши мне /reg')
    else:
        ws = wb[str(message.from_user.id)]
        bot.send_message(message.from_user.id, f"{ws['A1'].value}, что ты хочешь? "
                                               f"\n /day отчет за день "
                                               f"\n /month отчет за месяц"
                                               f"\n /add добавить покупочку"
                                               f"\n /category смотреть по категориям"
                                               f"\n /shop смотреть по магазинам")
        bot.register_next_step_handler(message, choose_action)


def choose_action(message):
    # Функия, которая распределяет активности пользователя по остальным функциям
    # (20 значущих строчек кода, границу проходят)
    er_id = message.from_user.id
    if message.text == '/day':
        need_date = str(datetime.datetime.now()).split(' ')[0]
        report = day_report(er_id, need_date)
        bot.send_message(message.from_user.id, report+f'\n что-нибудь еще? /menu')
    elif message.text == '/month':
        need_date = str(datetime.datetime.now()).split(' ')[0].split('-')[1]
        report = month_report(er_id, need_date)
        bot.send_message(er_id, report+f'\n что-нибудь еще? /menu')
    elif message.text == '/add':
        bot.send_message(er_id, f'Как вы хотите зарегистрировать покупку? \n'
                         + ' /category по категориям \n /shops по магазинам\n /anonim без привязки')
        bot.register_next_step_handler(message, purchase_value)
    elif message.text == '/category':
        bot.send_message(er_id, 'выберите промежуток\n /day за день\n /month за месяц')
        bot.register_next_step_handler(message, category_checker)
    elif message.text == '/shop':
        bot.send_message(er_id, 'выберите промежуток\n /day за день\n /month за месяц')
        bot.register_next_step_handler(message,shop_checker)
    else:
        bot.send_message(er_id, "я не знаю таких команд")


def category_checker(message):
    if message.text=='/day':
        categories = criteria_searcher('category', str(message.from_user.id))
        answer = 'Какую категорию вы хотите выбрать? '
        for x in categories:
            answer += f"{x} \n"
        bot.send_message(message.from_user.id, answer)
        bot.register_next_step_handler(message, category_report_day)
    elif message.text == '/month':
        categories = criteria_searcher('category', str(message.from_user.id))
        answer = 'Какую категорию вы хотите выбрать? '
        for x in categories:
            answer += f"{x} \n"
        bot.send_message(message.from_user.id, answer)
        bot.register_next_step_handler(message, category_report_month)
    else:
        bot.send_message(message.from_user.id,"эмм чегоо")


def category_report_month(message):
    categories = criteria_searcher('category', str(message.from_user.id))
    date = str(datetime.datetime.now()).split(' ')[0].split('-')[1]
    if message.text in categories:
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[str(message.from_user.id)]
        report = 0
        for i in range(2, ws.max_column + 1):
            if str(ws.cell(column=i, row=1).value).split(' ')[0].split('-')[1] == date and \
                    ws.cell(column=i, row=3).value == message.text:
                report += int(ws.cell(column=i, row=2).value)
        report = str(report)
        bot.send_message(message.from_user.id, report + f'\n что-нибудь еще? /menu')
    else:
        bot.send_message(message.from_user.id, " ничего не нашёл")

def category_report_day(message):
    categories = criteria_searcher('category', str(message.from_user.id))
    date = str(datetime.datetime.now()).split(' ')[0]
    if message.text in categories:
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[str(message.from_user.id)]
        report = 0
        for i in range(2, ws.max_column + 1):
            if str(ws.cell(column=i, row=1).value).split(' ')[0] == date and\
                ws.cell(column=i, row=3).value == message.text:
                report += int(ws.cell(column=i, row=2).value)
        report = str(report)
        bot.send_message(message.from_user.id, report+f'\n что-нибудь еще? /menu')
    else:
        bot.send_message(message.from_user.id, " ничего не нашёл")


def shop_checker():
    pass


def day_report(day_user_id, date):
    # Функция, которая принимает id юзера и возвращает его расходы за 1 день
    # (6 значущих строчек кода, границу проходят)
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    ws = wb[str(day_user_id)]
    report = 0
    for i in range(2, ws.max_column+1):
        if str(ws.cell(column=i, row=1).value).split(' ')[0] == date:
            report += int(ws.cell(column=i, row=2).value)
    return str(report)


def month_report(id_month_user, date):
    # Функция, которая принимает id юзера и возвращает его расходы за 1 месяц
    # (6 значущих строчек кода, границу проходят)
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    ws = wb[str(id_month_user)]
    report = 0
    for i in range(2, ws.max_column+1):
        if str(ws.cell(column=i, row=1).value).split(' ')[0].split('-')[1] == date:
            report += int(ws.cell(column=i, row=2).value)
    return str(report)


def purchase_value(message):
    # Функия, которая распределяет способы добавления расходов юзера по остальным функциям
    # (19 значущих строчек кода, границу проходят)
    if message.text == '/anonim':
        bot.send_message(message.from_user.id,
                         "Введите нужную сумму...")
        bot.register_next_step_handler(message, anonim_purchase)
    elif message.text == '/category':
        categories = criteria_searcher('category', str(message.from_user.id))
        answer = 'Какую категорию вы хотите выбрать? '\
                 + '(Возможные варианты представлены снизу)\n'
        for x in categories:
            answer += f"{x} \n"
        bot.send_message(message.from_user.id, answer)
        bot.register_next_step_handler(message, category_add)
    elif message.text == '/shops':
        shops = criteria_searcher('shop', str(message.from_user.id))
        answer = 'Какой магазин вы хотите выбрать? '\
                 + '(Возможные варианты представлены снизу)\n'
        for x in shops:
            answer += f"{x} \n"
        bot.send_message(message.from_user.id, answer)
        bot.register_next_step_handler(message, shop_add)


def criteria_searcher(criterion, user_id):
    criteria = set()
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    ws = wb[user_id]
    for i in range(2, ws.max_column + 1):
        if ws.cell(column=i, row=4).value == criterion:
            criteria.add(ws.cell(column=i, row=3).value)
    return criteria


def category_add(message):
    # фунция получает категорию в сообщении, записывает в глобальную переменную и запускает следующий шаг
    # (4 значущие строчки кода, границу проходят)
    global category
    category = message.text
    bot.send_message(message.from_user.id, "Введите нужную сумму...")
    bot.register_next_step_handler(message, category_purchase)


def category_purchase(message):
    # функция получает сумму, котрую хочет зарегистрировать человек,
    # записывает в файл сумму и категорию из глобальной переменой и возвращает в меню
    # (15 значущих строчек кода, границу проходят)
    global category
    category_id = str(message.from_user.id)
    try:
        value_category = int(message.text)
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[category_id]
        data = str(datetime.datetime.now()).split(' ')[0]
        ws.cell(column=ws.max_column + 1, row=1).value = data
        ws.cell(column=ws.max_column, row=2).value = value_category
        ws.cell(column=ws.max_column, row=3).value = category
        ws.cell(column=ws.max_column, row=4).value = 'category'
        category = ''
        wb.save('varmoney2.xlsx')
        bot.send_message(message.from_user.id, 'всё ок) \n /menu')
    except ValueError:
        bot.send_message(message.from_user.id, 'пиши цифрами, сука')


def shop_add(message):
    # фунция получает магазин в сообщении, записывает в глобальную переменную и запускает следующий шаг
    # (4 значущие строчки кода, границу проходят)
    global shop
    shop = message.text
    bot.send_message(message.from_user.id, "Введите нужную сумму...")
    bot.register_next_step_handler(message, shop_purchase)


def shop_purchase(message):
    # функция получает сумму, котрую хочет зарегистрировать человек,
    # записывает в файл сумму и магазин из глобальной переменной и возвращает в меню
    # (15 значущих строчек кода, границу проходят)
    global shop
    shop_id = str(message.from_user.id)
    try:
        shop_value = int(message.text)
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[shop_id]
        data = str(datetime.datetime.now()).split(' ')[0]
        ws.cell(column=ws.max_column + 1, row=1).value = data
        ws.cell(column=ws.max_column, row=2).value = shop_value
        ws.cell(column=ws.max_column, row=3).value = shop
        ws.cell(column=ws.max_column, row=4).value = 'shop'
        shop = ''
        wb.save('varmoney2.xlsx')
        bot.send_message(message.from_user.id, 'всё ок) \n /menu')
    except ValueError:
        bot.send_message(message.from_user.id, 'пиши цифрами, сука')


def anonim_purchase(message):
    # получает сумму и записывает в файл без категорий и магазинов
    # (11 значущих строчек кода, границу проходят)
    anonim_id = str(message.from_user.id)
    try:
        anonim_value = int(message.text)
        need_date = str(datetime.datetime.now()).split(' ')[0]
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[anonim_id]
        ws.cell(column=ws.max_column + 1, row=1).value = need_date
        ws.cell(column=ws.max_column, row=2).value = anonim_value
        wb.save('varmoney2.xlsx')
        bot.send_message(message.from_user.id, 'всё ок) \n /menu')
    except ValueError:
        bot.send_message(message.from_user.id, 'пиши цифрами, сука')


def get_name(message):
    # фунция получает ник человека в сообщении, записывает в глобальную переменную и запускает следующий шаг
    # (4 значущие строчки кода, границу проходят)
    global name
    name = message.text
    bot.send_message(message.from_user.id, "сколько тебе лет")
    bot.register_next_step_handler(message, get_age_and_id)


def get_age_and_id(message):
    # записывает в новую учетную запись ник, возраст человека и id его телеграмма
    # (18 значущих строчек кода, границу проходят)
    global age
    global name
    global id_user
    while age == 0:
        try:
            age = int(message.text)
        except ValueError:
            bot.send_message(message.from_user.id, 'пиши цифрами')
    id_for_age = str(message.from_user.id)
    wb = openpyxl.load_workbook('./varmoney2.xlsx')
    wb.create_sheet(id_for_age)
    sheet = wb[id_for_age]
    sheet['A1'] = name
    sheet['A2'] = age
    wb.save('varmoney2.xlsx')
    name = ''
    age = 0
    id_user = ''
    bot.send_message(message.from_user.id, 'cпaсибо за регистрацию \n /menu')


bot.polling(none_stop=True, interval=0)
